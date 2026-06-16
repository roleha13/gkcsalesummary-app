# processor.py

import os
import re
from datetime import datetime

import pdfplumber

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Alignment,
    PatternFill,
    Font,
    Border,
    Side
)

from openpyxl.chart import (
    LineChart,
    PieChart,
    BarChart,
    Reference
)

from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font



# =========================================================
# STYLES
# =========================================================

MONEY_FMT = '#,##0.00;(#,##0.00)'

LIGHT_ORANGE_FILL = PatternFill(
    start_color="FCE4D6",
    end_color="FCE4D6",
    fill_type="solid"
)

DARK_BLUE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F1F1F"
)

TOTAL_WINNINGS_FILL = PatternFill(
    start_color="F4B183",
    end_color="F4B183",
    fill_type="solid"
)

thin = Side(style='thin', color='000000')
thick = Side(style='thick', color='000000')
double = Side(style='double', color='000000')

# STANDARD BORDER
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# HEADER BORDER
header_border = Border(
    left=thick,
    right=thick,
    top=thick,
    bottom=thick
)

# BODY BORDER
body_border = Border(
    left=thick,
    right=thick,
    bottom=thin
)

# TOTAL BORDER
total_border = Border(
    left=thick,
    right=thick,
    top=thick,
    bottom=double
)


# =========================================================
# PDF EXTRACTION
# =========================================================

def get_last_number_from_line_with_keyword(
    text_block: str,
    keywords
):

    if isinstance(keywords, str):
        keywords = [keywords]

    cleaned_keywords = [
        k.replace(" ", "").upper()
        for k in keywords
    ]

    for line in text_block.splitlines():

        line_clean = (
            line
            .replace(" ", "")
            .upper()
        )

        for keyword_clean in cleaned_keywords:

            if keyword_clean in line_clean:

                nums = re.findall(r"[\d,]+", line)

                if not nums:
                    continue

                value = float(nums[-1].replace(",", ""))

                return (
                    -value
                    if '(' in line and ')' in line
                    else value
                )

    raise ValueError(
        f"Could not find numeric value for keywords: {keywords}"
    )


def extract_values_from_pdf(pdf_path: str):

    with pdfplumber.open(pdf_path) as pdf:

        text = "\n".join(
            page.extract_text() or ""
            for page in pdf.pages
        )

    date_match = re.search(
        r"Date\s+(\d{1,2}-[A-Za-z]{3}-\d{2})",
        text
    )

    if not date_match:
        raise ValueError(f"Date not found in {pdf_path}")

    date_obj = datetime.strptime(
        date_match.group(1),
        "%d-%b-%y"
    )

    ar_text = re.search(
        r"AMERICAN ROULETTE(.*?CARDS)",
        text,
        re.S
    ).group(1)

    cards_text = re.search(
        r"CARDS(.*?TABLES TOTALS)",
        text,
        re.S
    ).group(1)

    slots_text = re.search(
        r"SLOTS\s*\nBANK No\..*",
        text,
        re.S
    ).group(0)

    return {

        "date": date_obj,

        "TABLE AR":
            get_last_number_from_line_with_keyword(
                ar_text,
                "Sub-Totals"
            ),

        "TABLE CARDS":
            get_last_number_from_line_with_keyword(
                cards_text,
                "Sub-Totals"
            ),

        "SLOTS AT+CT+EGT":
            get_last_number_from_line_with_keyword(
                slots_text,
                [
                    "SLOTS AT+CT+EGT",
                    "SLOTS AT+CT"
                ]
            ),

        "SLOTS EG+AM+NOV":
            get_last_number_from_line_with_keyword(
                slots_text,
                "SLOTS EG+AM+NOV"
            ),

        "SLOTS TBJ":
            get_last_number_from_line_with_keyword(
                slots_text,
                "SLOTS TBJ"
            ),
    }


# =========================================================
# MAIN PROCESSING
# =========================================================

def process_pdfs_to_excel(pdf_files, output_folder):

    rows = [
        extract_values_from_pdf(pdf)
        for pdf in pdf_files
    ]

    rows.sort(key=lambda r: r['date'])

    first_date = rows[0]['date']

    month_name = first_date.strftime('%B').upper()
    month_abbrev_year = first_date.strftime('%b-%y')
    year_full = first_date.year
    month_number = first_date.month

    wb = Workbook()

    ws = wb.active
    ws.title = 'Sales Summary'

    # REMOVE GRIDLINES
    ws.sheet_view.showGridLines = False

    headers = [
        'Date',
        'TABLE AR',
        'TABLE CARDS',
        'SLOTS AT+CT+EGT',
        'SLOTS EG+AM+NOV',
        'SLOTS TBJ',
        'TOTAL WINNINGS',
        'TIPS',
        'GROSS INCOME'
    ]

    num_cols = len(headers)

    # =====================================================
    # MAIN TITLES
    # =====================================================

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=num_cols
    )

    ws['A1'] = 'GOLDEN KEY CASINO'

    ws['A1'].font = Font(
        name='Arial Black',
        size=12,
        bold=True
    )

    ws['A1'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=num_cols
    )

    ws['A2'] = 'SALES ANALYSIS FOR'

    ws['A2'].font = Font(
        name='Arial Black',
        size=12,
        bold=True
    )

    ws['A2'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=num_cols
    )

    ws['A3'] = month_abbrev_year

    ws['A3'].fill = LIGHT_ORANGE_FILL

    ws['A3'].font = Font(
        name='Arial Black',
        size=12,
        bold=True
    )

    ws['A3'].alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # =====================================================
    # HEADERS
    # =====================================================

    ws.append(headers)

    for cell in ws[4]:

        cell.font = Font(
            name='Arial',
            size=11,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        cell.border = header_border

    # =====================================================
    # DATA
    # =====================================================

    start_row = 5

    for i, row in enumerate(rows, start=start_row):

        # DATE COLUMN
        ws[f'A{i}'] = row['date']

        ws[f'A{i}'].number_format = 'd-mmm-yy'

        ws[f'A{i}'].fill = LIGHT_ORANGE_FILL

        vals = [
            'TABLE AR',
            'TABLE CARDS',
            'SLOTS AT+CT+EGT',
            'SLOTS EG+AM+NOV',
            'SLOTS TBJ'
        ]

        for col, key in enumerate(vals, start=2):

            ws.cell(
                row=i,
                column=col,
                value=row[key]
            ).number_format = MONEY_FMT

        # TOTAL WINNINGS
        ws[f'G{i}'] = f'=SUM(B{i}:F{i})'
        ws[f'G{i}'].number_format = MONEY_FMT
        ws[f'G{i}'].fill = TOTAL_WINNINGS_FILL


        # TIPS
        ws[f'H{i}'].number_format = MONEY_FMT

        # GROSS INCOME
        ws[f'I{i}'] = f'=G{i}+H{i}'
        ws[f'I{i}'].number_format = MONEY_FMT

        # BODY STYLING
        for col in range(1, num_cols + 1):

            cell = ws.cell(row=i, column=col)

            # keep bold for TOTAL WINNINGS column
            if col == 7:  # column G
                cell.font = Font(name='Arial', size=11, bold=True)
            else:
                cell.font = Font(name='Arial', size=11)

            cell.border = body_border

    # =====================================================
    # TOTAL ROW
    # =====================================================

    total_row = start_row + len(rows)

    ws[f'A{total_row}'] = 'TOTAL'

    ws[f'A{total_row}'].font = Font(
        name='Arial',
        size=11,
        bold=True
    )

    ws[f'A{total_row}'].fill = LIGHT_ORANGE_FILL

    ws[f'A{total_row}'].border = total_border

    for col_idx in range(2, num_cols + 1):

        col_letter = get_column_letter(col_idx)

        ws[f'{col_letter}{total_row}'] = (
            f'=SUM({col_letter}{start_row}:'
            f'{col_letter}{total_row-1})'
        )

        ws[f'{col_letter}{total_row}'].number_format = MONEY_FMT

        ws[f'{col_letter}{total_row}'].font = Font(
            name='Arial',
            size=11,
            bold=True
        )

        ws[f'{col_letter}{total_row}'].border = total_border

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    widths = {
        'A': 15,
        'B': 18,
        'C': 18,
        'D': 22,
        'E': 22,
        'F': 18,
        'G': 20,
        'H': 15,
        'I': 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # =====================================================
    # DASHBOARD SHEET
    # =====================================================

    dashboard = wb.create_sheet('Dashboard')

    dashboard.sheet_view.showGridLines = False

    dashboard.merge_cells('A1:H1')

    dashboard['A1'] = (
        'GOLDEN KEY CASINO – DASHBOARD TEMPLATE'
    )

    dashboard['A1'].font = Font(
        color='FFFFFF',
        bold=True,
        size=16
    )
    dashboard['A1'].fill = DARK_BLUE_FILL
    
    dashboard['A1'].alignment = Alignment(
        horizontal='center'
    )

    # =====================================================
    # SUMMARY TABLE
    # =====================================================

    dashboard['A3'] = 'Gaming Results Summary'

    dashboard['A3'].font = Font(
        bold=True,
        size=12
    )

    metrics = [
        'TABLE AR',
        'TABLE CARDS',
        'SLOTS AT+CT+EGT',
        'SLOTS EG+AM+NOV',
        'SLOTS TBJ'
    ]

    for idx, metric in enumerate(metrics, start=4):

        dashboard[f'A{idx}'] = metric

        dashboard[f'B{idx}'] = (
            f"='Sales Summary'!"
            f"{get_column_letter(idx-2)}{total_row}"
        )

        dashboard[f'B{idx}'].number_format = MONEY_FMT

        dashboard[f'A{idx}'].border = border
        dashboard[f'B{idx}'].border = border
        dashboard[f'C{idx}'].border = border

    dashboard['A9'] = 'TOTAL WINNINGS'

    dashboard['B9'] = (
        f"='Sales Summary'!G{total_row}"
    )

    dashboard['B9'].number_format = MONEY_FMT

    dashboard['A9'].border = border
    dashboard['B9'].border = border
    dashboard['C9'].border = border

    for idx in range(4, 9):

        dashboard[f'C{idx}'] = f'=B{idx}/$B$9'

        dashboard[f'C{idx}'].number_format = '0%'

    dashboard['C9'] = '=B9/B9'

    dashboard['C9'].number_format = '0%'

    dashboard['C9'].font = Font(bold=True)

    # =====================================================
    # PIE CHART
    # =====================================================

    pie = PieChart()

    pie.add_data(
        Reference(
            dashboard,
            min_col=2,
            min_row=4,
            max_row=8
        )
    )

    pie.set_categories(
        Reference(
            dashboard,
            min_col=1,
            min_row=4,
            max_row=8
        )
    )

    pie.title = 'Gaming Results Summary'

    pie.style = 10

    pie.height = 8
    pie.width = 10

    pie.legend.position = 'r'

    pie.dataLabels = DataLabelList()

    pie.dataLabels.showPercent = True
    pie.dataLabels.showLeaderLines = True

    slice_colors = [
        'C0504D',
        '4F81BD',
        '9BBB59',
        '8064A2',
        'F79646'
    ]

    for i, color in enumerate(slice_colors):

        pt = DataPoint(idx=i)

        pt.graphicalProperties.solidFill = color

        pie.series[0].data_points.append(pt)

    dashboard.add_chart(pie, 'D3')

    # =====================================================
    # CATEGORY DATES
    # =====================================================

    line_dates = Reference(
        ws,
        min_col=1,
        min_row=5,
        max_row=total_row-1
    )

    # =====================================================
    # TOTAL WINNINGS LINE CHART
    # =====================================================

    line = LineChart()

    data = Reference(
        ws,
        min_col=7,
        max_col=7,
        min_row=4,
        max_row=total_row-1
    )

    line.add_data(
        data,
        titles_from_data=True
    )

    line.set_categories(line_dates)

    line.title = 'Monthly Total Winnings'

    line.style = 13

    line.x_axis.title = 'Day of Month'
    line.y_axis.title = 'Amount'

    line.width = 18
    line.height = 8

    line.legend = None

    line.y_axis.majorGridlines = ChartLines()

    s1 = line.series[0]

    s1.marker.symbol = "circle"
    s1.marker.size = 7

    s1.graphicalProperties.line.width = 25000

    dashboard.add_chart(line, 'D30')

    # =====================================================
    # MONTHLY TIPS LINE CHART
    # =====================================================

    tips = LineChart()

    data = Reference(
        ws,
        min_col=8,
        max_col=8,
        min_row=4,
        max_row=total_row-1
    )

    tips.add_data(
        data,
        titles_from_data=True
    )

    tips.set_categories(line_dates)

    tips.title = 'Monthly Tips'

    tips.style = 13

    tips.x_axis.title = 'Day of Month'
    tips.y_axis.title = 'Amount'

    tips.width = 18
    tips.height = 8

    tips.legend = None

    tips.y_axis.majorGridlines = ChartLines()

    s1 = tips.series[0]

    s1.marker.symbol = "circle"
    s1.marker.size = 7

    s1.graphicalProperties.line.width = 25000

    dashboard.add_chart(tips, 'D50')

    # =====================================================
    # STACKED BAR CHART
    # =====================================================

    stacked = BarChart()

    data = Reference(
        ws,
        min_col=2,
        max_col=6,
        min_row=4,
        max_row=total_row-1
    )

    stacked.add_data(
        data,
        titles_from_data=True,
        from_rows=False
    )

    stacked.set_categories(line_dates)

    stacked.type = 'col'

    stacked.grouping = 'stacked'

    stacked.overlap = 100

    stacked.title = 'Daily Gaming Mix'

    stacked.style = 12

    stacked.x_axis.title = 'Day of Month'
    stacked.y_axis.title = 'Amount'

    stacked.width = 20
    stacked.height = 10

    stacked.legend.position = 'r'

    stacked.y_axis.majorGridlines = ChartLines()

    stack_colors = [
        '4F81BD',
        'C0504D',
        '9BBB59',
        '8064A2',
        'F2C811'
    ]

    for idx, series in enumerate(stacked.series):

        series.graphicalProperties.solidFill = stack_colors[idx]

    dashboard.add_chart(stacked, 'D75')

    # =====================================================
    # SAVE FILE
    # =====================================================

    file_name = (
        f"{month_number:02d}. SALES ANALYSIS "
        f"{month_name} {year_full}.xlsx"
    )

    output_path = os.path.join(
        output_folder,
        file_name
    )

    wb.save(output_path)

    return output_path
